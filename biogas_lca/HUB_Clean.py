import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Load the Excel file
input_file_path = r"C:\Users\ColinY\HUB REPORT.xlsx"
output_file_path = r"C:\Users\ColinY\HUB REPORT_Cleaned.xlsx"
df = pd.read_excel(input_file_path)

# Split the 'start_time' column into 'Date' and 'Time'
df['Date'] = pd.to_datetime(df['start_time']).dt.date
df['Time'] = pd.to_datetime(df['start_time']).dt.time

# Drop the 'start_time' and 'end_time' columns
df = df.drop(columns=['start_time', 'end_time'])

# Reorder columns to have 'Date' and 'Time' first
columns_order = ['Date', 'Time'] + [col for col in df.columns if col not in ['Date', 'Time']]
df = df[columns_order]

# Save the cleaned data to a new Excel file
df.to_excel(output_file_path, index=False)

# Load the workbook and add an Excel table
wb = load_workbook(output_file_path)
ws = wb.active

# Define the table range
last_column = get_column_letter(len(df.columns))  # Get the last column letter
table_range = f"A1:{last_column}{len(df) + 1}"  # Adjust for column count and row count
table = Table(displayName="CleanedData", ref=table_range)

# Apply a table style
style = TableStyleInfo(
    name="TableStyleMedium9", 
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True,
)
table.tableStyleInfo = style
ws.add_table(table)

# Save the workbook with the formatted table
wb.save(output_file_path)

# Display a success message
print(f"Cleaned file saved to {output_file_path} with an Excel table.")
