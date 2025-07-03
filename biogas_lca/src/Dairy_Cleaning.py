import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

input_path = r"C:\Users\XXX\DAIRY DATA.xlsx"
output_path = r"C:\Users\xxx\DAIRY DATA_cleaned.xlsx" 
log_file_path = r"C:\Users\xxx\change_log.txt"

df = pd.read_excel(input_path, engine="openpyxl")
df_original = df.copy()  # For logging original vs. cleaned

# prepare data format change
if "Time" in df.columns:
    df["Time"] = pd.to_datetime(df["Time"], errors="coerce")
    df["DateTime"] = df["Time"]  # Create a DateTime column that retains both date and time
else:
    df.rename(columns={df.columns[0]: "DateTime"}, inplace=True)
    
df["DateTime"] = pd.to_datetime(df["DateTime"], errors="coerce")

# Debug: Check the DateTime column
print("\nDateTime column (before separation):")
print(df["DateTime"].head())

# Separate DateTime into Date and Time columns
df["Date"] = df["DateTime"].dt.date  # Extract the date part
df["Time"] = df["DateTime"].dt.time  # Extract the time part

# Debug: Print the first few rows to verify
print("\nSeparated Date and Time columns:")
print(df[["DateTime", "Date", "Time"]].head())

# Reorder columns: Move 'Date' to the first column, followed by 'Time', and then everything else
columns_order = ['Date', 'Time'] + [col for col in df.columns if col not in ['Date', 'Time']]
df = df[columns_order]

flow_columns = [c for c in df.columns if "TF_TOTAL_VOLUME_ACCUM" in c or "TF_TOTAL_RECIRC_ACCUM" in c]
methane_columns = [c for c in df.columns if "TF_METHANE" in c]

negative_value_changes = {col: set() for col in flow_columns}
forward_fill_changes   = {col: set() for col in flow_columns}
no_decreasing_changes  = {col: set() for col in flow_columns}
methane_zero_replaced  = {col: set() for col in methane_columns}

# Cleaning Data with forward fill, no decreasing, fill remaining NaN with 0. 
for col in flow_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # (a) Replace any negative values with 0 and track them.
    for i in range(len(df)):
        if pd.notna(df.at[i, col]) and df.at[i, col] < 0:
            df.at[i, col] = 0
            negative_value_changes[col].add(i)
    
    # (b) Manual forward-fill for missing values
    for i in range(len(df)):
        if pd.isna(df.at[i, col]) and i > 0:
            df.at[i, col] = df.at[i - 1, col]
            forward_fill_changes[col].add(i)
    
    # (c) Prevent any decrease (if current < previous, replace it)
    for i in range(1, len(df)):
        if pd.notna(df.at[i, col]) and pd.notna(df.at[i - 1, col]):
            # Only adjust if the decrease is more than 0.01
            if df.at[i, col] < df.at[i - 1, col] - 0.01:
                df.at[i, col] = df.at[i - 1, col]
                no_decreasing_changes[col].add(i)
    
    # (d) Finally, if any cells are still NaN (e.g., first row might be NaN), fill them with 0.
    for i in range(len(df)):
        if pd.isna(df.at[i, col]):
            df.at[i, col] = 0
            negative_value_changes[col].add(i)


# REPLACE 0 OR NaN WITH DAILY AVG
df["DateTime"] = pd.to_datetime(df["DateTime"], errors="coerce")
unique_dates = sorted(df["DateTime"].dropna().unique())

for col in methane_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce")
    df[col] = df[col].ffill()  # Forward-fill missing methane values

    # (a) Compute daily average ignoring zeros.
    daily_avg = {}
    for date_val in unique_dates:
        subset = df.loc[df["DateTime"] == date_val, col]
        non_zero = subset[subset > 0].dropna()
        daily_avg[date_val] = non_zero.mean() if len(non_zero) > 0 else None

    # (b) For days with no non-zero values, fill with previous day's average.
    for i, d in enumerate(unique_dates):
        if daily_avg[d] is None and i > 0:
            daily_avg[d] = daily_avg[unique_dates[i - 1]]
    
    # (c) Replace 0 or NaN with the daily average and track these changes.
    for i in range(len(df)):
        row_date = df.at[i, "DateTime"]
        if pd.isna(row_date):
            continue
        if row_date in daily_avg and daily_avg[row_date] is not None:
            old_val = df.at[i, col]
            if pd.isna(old_val) or old_val == 0:
                df.at[i, col] = daily_avg[row_date]
                methane_zero_replaced[col].add(i)

# Divide all TF_METHANE values by 100
for col in methane_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce") / 100

# dynamic calculation change for farms with recirculation
recirc_farms = {
    "02_xxx",
    "01_xxx",
    "06_xxx",
    "07_xxx",
    "03_xxx",
    "04_xxx",
    "05_xxx"
}

df_result = pd.DataFrame()
df_result["DateTime"] = df["DateTime"]

def parse_farm_name(methane_col):
    parts = methane_col.split("\\")
    return parts[1] if len(parts) >= 2 else "UnknownFarm"

num_cols = len(df.columns)
i = 1

while i + 2 < num_cols:
    methane_col = df.columns[i]
    accum_col   = df.columns[i + 1]
    recirc_col  = df.columns[i + 2]
    
    if ("TF_METHANE" in methane_col and
        "TF_TOTAL_VOLUME_ACCUM" in accum_col and
        "TF_TOTAL_RECIRC_ACCUM" in recirc_col):
        
        farm_name = parse_farm_name(methane_col)
        net_col_name = f"{farm_name} NET MSCF"
        
        df[accum_col]  = pd.to_numeric(df[accum_col], errors="coerce")
        df[recirc_col] = pd.to_numeric(df[recirc_col], errors="coerce")
        
        if farm_name in recirc_farms:
            df_result[net_col_name] = (df[accum_col].diff() - df[recirc_col].diff()).abs()
        else:
            df_result[net_col_name] = df[accum_col].diff().abs()
        
        i += 3
    else:
        i += 1

# Output to Excel

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_result.to_excel(writer, sheet_name="NET MSCF", index=False)
    
    # Ensure 'Date' is the first column, followed by 'Time', and then all other columns
    columns_order = ['Date', 'Time'] + [col for col in df.columns if col not in ['Date', 'Time']]
    df = df[columns_order]

    # Write the reordered DataFrame to the "Cleaned Data" sheet
    df.to_excel(writer, sheet_name="Cleaned Data", index=False)


# Convet Sheet to Excel Table
wb = load_workbook(output_path)

# Convert "NET MSCF" sheet to a table
ws_result = wb["NET MSCF"]
table_result = Table(displayName="NetMSCFTable", ref=ws_result.dimensions)
style_result = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
)
table_result.tableStyleInfo = style_result
ws_result.add_table(table_result)

# Convert "Cleaned Data" sheet to a table
ws_cleaned = wb["Cleaned Data"]
table_cleaned = Table(displayName="CleanedDataTable", ref=ws_cleaned.dimensions)
style_cleaned = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
)
table_cleaned.tableStyleInfo = style_cleaned
ws_cleaned.add_table(table_cleaned)

# Save the workbook
wb.save(output_path)


# Excel colored changed cell for different catagories

wb = load_workbook(output_path)
ws_cleaned = wb["Cleaned Data"]

# Define fill colors
blue_fill   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Negative/missing (set to 0)
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Forward-fill
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # No-decreasing adjustments
red_fill    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Methane replacements

# Highlight negative or missing values replaced with 0 (blue)
for col in flow_columns:
    col_idx = df.columns.get_loc(col) + 1
    for row_idx in negative_value_changes[col]:
        ws_cleaned.cell(row=row_idx + 2, column=col_idx).fill = blue_fill

# Highlight forward-fill changes (orange)
for col in flow_columns:
    col_idx = df.columns.get_loc(col) + 1
    for row_idx in forward_fill_changes[col]:
        ws_cleaned.cell(row=row_idx + 2, column=col_idx).fill = orange_fill

# Highlight no-decreasing changes (yellow)
for col in flow_columns:
    col_idx = df.columns.get_loc(col) + 1
    for row_idx in no_decreasing_changes[col]:
        ws_cleaned.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill

# Highlight methane replacements (red)
for col in methane_columns:
    col_idx = df.columns.get_loc(col) + 1
    for row_idx in methane_zero_replaced[col]:
        ws_cleaned.cell(row=row_idx + 2, column=col_idx).fill = red_fill

wb.save(output_path)


# LOG ALL CHANGES TO A TEXT FILE
with open(log_file_path, "w", encoding="utf-8") as log_file:
    log_file.write("==== Negative or Missing Flow Values Replaced with 0 (Blue) ====\n")
    for col in flow_columns:
        for row_idx in sorted(negative_value_changes[col]):
            old_val = df_original.at[row_idx, col]
            log_file.write(
                f"Row {row_idx}, Column '{col}': Negative/missing value {old_val:.3f} replaced with 0\n"
                if pd.notna(old_val) else
                f"Row {row_idx}, Column '{col}': Negative/missing value NaN replaced with 0\n"
            )

    log_file.write("\n==== Forward-Fill Changes (Orange) ====\n")
    for col in flow_columns:
        for row_idx in sorted(forward_fill_changes[col]):
            old_val = df_original.at[row_idx, col]
            new_val = df.at[row_idx, col]
            log_file.write(
                f"Row {row_idx}, Column '{col}': was {old_val:.3f}, forward-filled to {new_val:.3f}\n"
                if pd.notna(old_val) and pd.notna(new_val) else
                f"Row {row_idx}, Column '{col}': was {old_val}, forward-filled to {new_val}\n"
            )

    log_file.write("\n==== No-Decreasing Changes (Yellow) ====\n")
    for col in flow_columns:
        for row_idx in sorted(no_decreasing_changes[col]):
            old_val = df_original.at[row_idx, col]
            new_val = df.at[row_idx, col]
            log_file.write(
                f"Row {row_idx}, Column '{col}': was {old_val:.3f}, adjusted to {new_val:.3f} to prevent decrease\n"
                if pd.notna(old_val) and pd.notna(new_val) else
                f"Row {row_idx}, Column '{col}': was {old_val}, adjusted to {new_val} to prevent decrease\n"
            )

    log_file.write("\n==== Methane Zero Replacements (Red) ====\n")
    for col in methane_columns:
        for row_idx in sorted(methane_zero_replaced[col]):
            old_val = df_original.at[row_idx, col]
            new_val = df.at[row_idx, col]
            log_file.write(
                f"Row {row_idx}, Column '{col}': was {old_val:.3f}, replaced with daily average {new_val:.3f}\n"
                if pd.notna(old_val) and pd.notna(new_val) else
                f"Row {row_idx}, Column '{col}': was {old_val}, replaced with daily average {new_val}\n"
            )

print("\nDone! Cleaned NET MSCF data saved to:", output_path)
print("Change log saved to:", log_file_path)
print("Sheets:")
print(" - NET MSCF (final calculations, dynamic farms)")
print(" - Cleaned Data (flow fixes & methane replaced with daily avg)")
